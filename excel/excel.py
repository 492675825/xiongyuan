import openpyxl

#导入excel文件
wb = openpyxl.load_workbook(r'output.xlsx')

#getting sheets from the workbook

#print sheetnames
print(wb.sheetnames)

for sheet in wb:
    #print sheet name
    print(sheet.title)

#生成一个sheet，名叫mySheet
mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

#获取某个sheet的内容
sheet3 = wb['Sheet1']

#获取单元格中某个单元格的值
ws = wb['Sheet1']
print(ws['A1'].value)
print(ws['A2'].value)




